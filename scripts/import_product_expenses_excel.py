#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook


def compose_key(*parts: Any) -> str:
    values = [str(part or "").strip() for part in parts]
    if any(not value for value in values):
        return ""
    return "||".join(values)


def load_rows(xlsx_path: Path, sheet_name: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    headers = list(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    rows: List[Dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        record = {str(header): row[idx] for idx, header in enumerate(headers)}
        rows.append(record)
    return headers, rows


def build_mapping(rows: List[Dict[str, Any]]) -> Tuple[Dict[str, float], List[str]]:
    by_label: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in rows:
        label = str(row.get("item_label") or "").strip()
        if not label:
            continue
        by_label[label].append({
            "label": label,
            "sku": str(row.get("product_sku") or "").strip(),
            "import_code": str(row.get("item_import_code") or "").strip(),
            "warehouse_number": str(row.get("item_warehouse_number") or "").strip(),
            "cost": float(row.get("current_expense_per_item_net") or 0),
        })

    mapping: Dict[str, float] = {}
    unresolved: List[str] = []

    for label, entries in by_label.items():
        unique_costs = sorted({round(entry["cost"], 6) for entry in entries})
        if len(unique_costs) == 1:
            mapping[label] = float(entries[0]["cost"])
            continue

        default_entries = [
            entry for entry in entries
            if not entry["warehouse_number"] and not entry["import_code"]
        ]
        if len({round(entry["cost"], 6) for entry in default_entries}) == 1 and default_entries:
            mapping[label] = float(default_entries[0]["cost"])

        added_any = False
        for entry in entries:
            for candidate_key in (
                compose_key(label, entry["warehouse_number"]),
                compose_key(label, entry["import_code"]),
                compose_key(label, entry["sku"]),
            ):
                if candidate_key:
                    mapping[candidate_key] = float(entry["cost"])
                    added_any = True
                    break

        if not added_any:
            unresolved.append(label)

    return mapping, sorted(set(unresolved))


def main() -> int:
    parser = argparse.ArgumentParser(description="Import product expenses from edited Excel workbook.")
    parser.add_argument("--xlsx", required=True, help="Path to edited workbook")
    parser.add_argument("--sheet", default="All products", help="Worksheet name")
    parser.add_argument("--output", required=True, help="Output JSON mapping path")
    parser.add_argument(
        "--keep-json",
        default="",
        help="Optional existing JSON path to merge/keep extra keys from (e.g. zero-cost service items)",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    output_path = Path(args.output)
    _, rows = load_rows(xlsx_path, args.sheet)
    mapping, unresolved = build_mapping(rows)

    if args.keep_json:
        keep_path = Path(args.keep_json)
        if keep_path.exists():
            with keep_path.open("r", encoding="utf-8") as handle:
                existing = json.load(handle) or {}
            for key, value in existing.items():
                if key not in mapping:
                    mapping[str(key)] = float(value)

    output_path.write_text(
        json.dumps(dict(sorted(mapping.items(), key=lambda item: item[0].lower())), ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(f"Imported {len(rows)} Excel rows into {output_path}")
    print(f"Generated {len(mapping)} expense keys")
    if unresolved:
        print("Unresolved labels:")
        for label in unresolved:
            print(f"- {label}")
    else:
        print("No unresolved labels")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
